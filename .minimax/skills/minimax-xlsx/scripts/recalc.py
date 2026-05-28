"""xlsx 公式回算工具：基于 LibreOffice 与 openpyxl，输出 JSON 错误清单。

Part of the MiniMax xlsx skill (MIT). See LICENSE for terms.

CLI 形态保持与上游契约一致：

    python scripts/recalc.py <workbook> [timeout_seconds]

输出 JSON 字段固定为 ``status`` / ``total_errors`` / ``total_formulas`` /
``error_summary``；错误路径单返 ``{"error": "..."}``。
"""

from __future__ import annotations

import argparse
import json
import os
import platform
import signal
import subprocess
import sys
from pathlib import Path

from office.soffice import get_soffice_env
from openpyxl import load_workbook

# --- LibreOffice Basic 模块落盘位置 -----------------------------------------
BASIC_DIR_DARWIN = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
BASIC_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
BASIC_MODULE_FILE = "Module1.xba"

# 注：宏内容与 XML 必须保持上游字面值，否则 LibreOffice 找不到 RecalculateAndSave。
LO_RECALC_BASIC = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

# 与 SKILL.md 一致：按字母序列出 7 类常见 Excel 错误标记。
XLSX_ERROR_TOKENS = (
    "#DIV/0!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#VALUE!",
)

_BASIC_INVOCATION = (
    "vnd.sun.star.script:Standard.Module1.RecalculateAndSave"
    "?language=Basic&location=application"
)
_MAX_LOCATIONS_PER_TOKEN = 20


def _openpyxl_compat_hint(exc: Exception | str) -> str:
    """Return a user-facing hint for known openpyxl compatibility failures."""
    msg = str(exc)
    low = msg.lower()
    if "expected <class 'int'>" in msg or "merge" in low or "merged" in low:
        return (
            "openpyxl failed while parsing merged-cell metadata after LibreOffice "
            "saved the workbook. This is a known non-standard/rewritten mergeCell "
            "compatibility issue; recalc.py will use raw XML scanning, so do not "
            "retry blindly. If you need to read values later, use "
            "load_workbook(..., read_only=True, data_only=True) or raw XML."
        )
    if "chart" in low or "drawing" in low or "graphic" in low:
        return (
            "openpyxl failed while parsing chart/drawing metadata. Formula/error "
            "verification does not need chart objects, so recalc.py will bypass "
            "openpyxl and scan worksheet XML directly. Do not retry unless the "
            "raw-XML fallback also fails."
        )
    if "stylesheet" in low or "style" in low or "numberformat" in low:
        return (
            "openpyxl failed while parsing styles/number formats. Formula/error "
            "verification can continue via raw worksheet XML, but downstream code "
            "should avoid rewriting this workbook with openpyxl unless style "
            "round-trip has been tested."
        )
    return (
        "openpyxl could not parse this workbook, likely due to non-standard XLSX "
        "metadata. recalc.py will try a raw-XML scan that ignores charts, drawings, "
        "styles, and merged-cell objects and only verifies worksheet formulas/errors."
    )


# --- subprocess helpers ------------------------------------------------------
def has_coreutils_timeout() -> bool:
    """macOS 下 coreutils 提供 ``gtimeout``；Linux 用内建 ``timeout``。"""
    try:
        subprocess.run(
            ["gtimeout", "--version"],
            capture_output=True,
            timeout=1,
            check=False,
        )
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def _wrap_with_timeout(cmd: list[str], timeout_sec: int) -> list[str]:
    sysname = platform.system()
    if sysname == "Linux":
        return ["timeout", str(timeout_sec), *cmd]
    if sysname == "Darwin" and has_coreutils_timeout():
        return ["gtimeout", str(timeout_sec), *cmd]
    return cmd


def _build_soffice_cmdline(workbook_path: str) -> list[str]:
    return [
        "soffice",
        "--headless",
        "--norestore",
        _BASIC_INVOCATION,
        workbook_path,
    ]


# --- LibreOffice Basic bootstrap --------------------------------------------
def install_recalc_basic() -> bool:
    """把 RecalculateAndSave 宏写入 LibreOffice user profile；幂等。"""
    sysname = platform.system()
    basic_dir = os.path.expanduser(
        BASIC_DIR_DARWIN if sysname == "Darwin" else BASIC_DIR_LINUX
    )
    module_path = Path(basic_dir) / BASIC_MODULE_FILE

    # 已写入并仍包含目标 Sub —— 视作就绪
    if module_path.exists() and "RecalculateAndSave" in module_path.read_text():
        return True

    # user profile 还未生成时，先冷启一次 soffice 让它建好目录
    if not module_path.parent.exists():
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=10,
            env=get_soffice_env(),
        )
        module_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        module_path.write_text(LO_RECALC_BASIC)
        return True
    except Exception:
        return False


# --- 主流程 -------------------------------------------------------------------
def recalculate_workbook(
    workbook_path: str | os.PathLike,
    *,
    timeout_sec: int = 30,
) -> dict:
    """回算 workbook 内全部公式，返回上游约定 schema 的 JSON 字典。"""
    target = Path(workbook_path)
    if not target.exists():
        return {"error": f"workbook 不存在：{target}"}

    abs_path = str(target.resolve())

    if not install_recalc_basic():
        return {"error": "无法写入 LibreOffice Basic 回算模块"}

    cmd = _wrap_with_timeout(_build_soffice_cmdline(abs_path), timeout_sec)
    # Popen + start_new_session so the Python-side timeout fallback can
    # kill the whole process group. Without it, only the wrapper dies and
    # the LibreOffice child keeps the single-instance profile lock,
    # silently breaking every subsequent invocation.
    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        env=get_soffice_env(),
        start_new_session=True,
    )
    try:
        _stdout, stderr_text = proc.communicate(timeout=timeout_sec + 5)
    except subprocess.TimeoutExpired:
        try:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
        except (ProcessLookupError, PermissionError):
            proc.kill()
        proc.wait()
        return {
            "error": (
                f"soffice exceeded {timeout_sec}s timeout; killed entire "
                "process group via Python-side fallback. On macOS install "
                "coreutils (`brew install coreutils`) so the gtimeout "
                "wrapper can cap soffice cleanly."
            )
        }

    # 124 是 timeout 的退出码——文档已写回磁盘，仍可继续扫错
    if proc.returncode not in (0, 124):
        stderr_text = stderr_text or "Unknown error during recalculation"
        if "Module1" in stderr_text or "RecalculateAndSave" not in stderr_text:
            return {"error": "LibreOffice Basic 模块未正确加载"}
        return {"error": stderr_text}

    # LibreOffice succeeded; openpyxl may still choke on rewritten merged
    # cells. Try the openpyxl path first, fall back to a raw-XML scan that
    # bypasses merged-cell parsing entirely.
    try:
        return _scan_for_errors(target)
    except Exception as exc:
        # openpyxl is strict about the full workbook package (merged cells,
        # chart drawings, styles, relationship targets). Recalc verification only
        # needs worksheet cached values and <f> formula nodes, so fall back to a
        # raw-XML scan for ANY openpyxl parse failure instead of crashing or
        # encouraging blind retries.
        try:
            return _scan_for_errors_via_xml(target, exc)
        except Exception as fallback_exc:
            return {
                "error": (
                    "LibreOffice recalc completed, but workbook verification failed. "
                    f"openpyxl_error={type(exc).__name__}: {exc}; "
                    f"xml_fallback_error={type(fallback_exc).__name__}: {fallback_exc}. "
                    "This is usually a non-standard XLSX package (merged cells, "
                    "charts/drawings, styles, or relationships). Do not retry the "
                    "same command blindly; inspect the workbook with "
                    "scripts/office/unpack.py or use Excel/LibreOffice to resave it."
                ),
                "openpyxl_error": str(exc),
                "openpyxl_error_type": type(exc).__name__,
                "xml_fallback_error": str(fallback_exc),
                "xml_fallback_error_type": type(fallback_exc).__name__,
                "compatibility_hint": _openpyxl_compat_hint(exc),
            }


def _scan_for_errors(target: Path) -> dict:
    """读两次 workbook：一次取计算后值找错误，一次原始公式数公式。"""
    findings: dict[str, list[str]] = {token: [] for token in XLSX_ERROR_TOKENS}
    error_count = 0

    book = load_workbook(target, data_only=True)
    try:
        for sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            for row in sheet.iter_rows():
                for node in row:
                    value = node.value
                    if not isinstance(value, str):
                        continue
                    for token in XLSX_ERROR_TOKENS:
                        if token in value:
                            findings[token].append(
                                f"{sheet_name}!{node.coordinate}"
                            )
                            error_count += 1
                            break
    finally:
        book.close()

    summary: dict[str, dict] = {}
    for token, places in findings.items():
        if places:
            summary[token] = {
                "count": len(places),
                "locations": places[:_MAX_LOCATIONS_PER_TOKEN],
            }

    formula_total = _count_formulas(target)

    return {
        "status": "success" if error_count == 0 else "errors_found",
        "total_errors": error_count,
        "error_summary": summary,
        "total_formulas": formula_total,
    }


def _count_formulas(target: Path) -> int:
    book = load_workbook(target, data_only=False)
    try:
        formula_total = 0
        for sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            for row in sheet.iter_rows():
                for node in row:
                    cv = node.value
                    if isinstance(cv, str) and cv.startswith("="):
                        formula_total += 1
        return formula_total
    finally:
        book.close()


# --- raw-XML fallback --------------------------------------------------------
# 触发场景：LibreOffice 把原文件里少量 mergeCell 改写成数十倍体量后，
# openpyxl 解析新生成的 ref 字段时拿到 ``min_col == None``，抛
# ``TypeError: expected <class 'int'>``。这个 fallback 不依赖 openpyxl，
# 直接用 stdlib 走 zipfile + ElementTree 完成同样的统计任务。
_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_DOC_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _scan_for_errors_via_xml(target: Path, openpyxl_error: Exception | str) -> dict:
    """Pure-stdlib equivalent of :func:`_scan_for_errors`.

    Used when LibreOffice rewrites a workbook in a form openpyxl cannot
    re-read (typically merged-cell ``ref`` mutations that yield ``min_col``
    of ``None``). Walks ``xl/worksheets/sheet*.xml`` directly, counts
    ``<f>`` elements as formulas, and looks for the seven Excel error
    tokens inside cached ``<v>`` values. Returns the same JSON schema as
    the openpyxl path with an extra ``warning`` field so callers can tell
    which scanner produced the verdict.
    """
    import zipfile
    from xml.etree import ElementTree as ET

    findings: dict[str, list[str]] = {token: [] for token in XLSX_ERROR_TOKENS}
    formula_total = 0

    with zipfile.ZipFile(target, "r") as zf:
        sheet_display = _build_sheet_display_map(zf, ET)

        for entry in sorted(zf.namelist()):
            if not entry.startswith("xl/worksheets/sheet") or not entry.endswith(".xml"):
                continue
            display = sheet_display.get(entry, entry.split("/")[-1])
            tree = ET.fromstring(zf.read(entry))
            for cell in tree.iter(f"{{{_NS_MAIN}}}c"):
                if cell.find(f"{{{_NS_MAIN}}}f") is not None:
                    formula_total += 1
                v_node = cell.find(f"{{{_NS_MAIN}}}v")
                if v_node is None or not v_node.text:
                    continue
                vtext = v_node.text.strip()
                for token in XLSX_ERROR_TOKENS:
                    if vtext.startswith(token):
                        findings[token].append(
                            f"{display}!{cell.get('r', '?')}"
                        )
                        break

    error_count = sum(len(places) for places in findings.values())
    summary: dict[str, dict] = {
        token: {
            "count": len(places),
            "locations": places[:_MAX_LOCATIONS_PER_TOKEN],
        }
        for token, places in findings.items()
        if places
    }

    return {
        "status": "success" if error_count == 0 else "errors_found",
        "total_errors": error_count,
        "error_summary": summary,
        "total_formulas": formula_total,
        "warning": (
            "openpyxl could not parse the post-recalc workbook "
            f"({type(openpyxl_error).__name__ if isinstance(openpyxl_error, Exception) else 'Error'}: {openpyxl_error}); "
            "verified via raw XML scan instead. LibreOffice did complete the recalculation."
        ),
        "scanner": "raw_xml_fallback",
        "openpyxl_error_type": type(openpyxl_error).__name__ if isinstance(openpyxl_error, Exception) else "Error",
        "openpyxl_error": str(openpyxl_error),
        "compatibility_hint": _openpyxl_compat_hint(openpyxl_error),
    }


def _build_sheet_display_map(zf, et_module) -> dict[str, str]:
    """Map ``xl/worksheets/sheetN.xml`` → user-visible sheet name.

    Returns the underlying entry path keyed by display name so the JSON
    locations match what reviewers see in Excel rather than zip internals.
    """
    wb_root = et_module.fromstring(zf.read("xl/workbook.xml"))
    rels_root = et_module.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    rid_to_target: dict[str, str] = {}
    for rel in rels_root.findall(f"{{{_NS_PKG_REL}}}Relationship"):
        rid_to_target[rel.get("Id", "")] = rel.get("Target", "")

    display: dict[str, str] = {}
    for sheet_node in wb_root.findall(
        f"{{{_NS_MAIN}}}sheets/{{{_NS_MAIN}}}sheet"
    ):
        rid = sheet_node.get(f"{{{_NS_DOC_REL}}}id", "")
        target = rid_to_target.get(rid, "").lstrip("./").lstrip("/")
        if not target:
            continue
        entry = target if target.startswith("xl/") else f"xl/{target}"
        display[entry] = sheet_node.get("name") or entry
    return display


# --- 向后兼容包装：旧外部调用 ``from recalc import recalc`` 仍可用 -----------
def recalc(filename, timeout=30):
    """Compatibility shim around :func:`recalculate_workbook`."""
    return recalculate_workbook(filename, timeout_sec=timeout)


# --- CLI 入口 ----------------------------------------------------------------
def _build_argparser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="recalc.py",
        description=(
            "Recalculate every formula in an .xlsx workbook through LibreOffice "
            "and emit a JSON report of remaining error markers."
        ),
        epilog=(
            "Output JSON keys: status, total_errors, total_formulas, error_summary; "
            "or {\"error\": ...} on failure. Tokens scanned: "
            + ", ".join(XLSX_ERROR_TOKENS)
            + "."
        ),
    )
    parser.add_argument(
        "workbook_path",
        help="Path to the .xlsx (or .xlsm) file to recalculate.",
    )
    parser.add_argument(
        "timeout_seconds",
        nargs="?",
        type=int,
        default=30,
        help="LibreOffice subprocess timeout in seconds (default: 30).",
    )
    return parser


def main(argv: list[str] | None = None) -> None:
    parser = _build_argparser()
    args = parser.parse_args(argv)
    payload = recalculate_workbook(
        args.workbook_path, timeout_sec=args.timeout_seconds
    )
    print(json.dumps(payload, indent=2))
    if "error" in payload and payload.get("status") is None:
        sys.exit(2)


if __name__ == "__main__":
    main()
